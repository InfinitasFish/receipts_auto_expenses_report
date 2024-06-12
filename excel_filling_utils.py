import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, cols_from_range, range_boundaries
from copy import copy, deepcopy
import numpy as np
import cv2
from PIL import Image


def parse_extracted_data(extracted_data):
    # extracted_data is list of ( ()-TOKEN, WORD ) tuples

    # collect all sequences
    seqs_dict = {'company_seqs': [], 'address_seqs': [], 'date_seqs': [], 'total_seqs': []}

    curr_address = []
    curr_company = []
    curr_date = []
    curr_total = []

    # ugly
    for i in range(len(extracted_data)):
        # check for entity name
        if 'ADDRESS' in extracted_data[i][0]:
            # different cases for I-B-I-I, B-I-I, I-I-I, B-I-B-I-I
            if 'B' in extracted_data[i][0] and curr_address == []:
                curr_address = [extracted_data[i][1]]
            elif 'B' in extracted_data[i][0] and curr_address != []:
                seqs_dict['address_seqs'].append(curr_address)
                curr_address = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_address == []:
                curr_address = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_address != []:
                curr_address.append(extracted_data[i][1])

        elif 'COMPANY' in extracted_data[i][0]:
            if 'B' in extracted_data[i][0] and curr_company == []:
                curr_company = [extracted_data[i][1]]
            elif 'B' in extracted_data[i][0] and curr_company != []:
                seqs_dict['company_seqs'].append(curr_company)
                curr_company = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_company == []:
                curr_company = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_company != []:
                curr_company.append(extracted_data[i][1])

        elif 'DATE' in extracted_data[i][0]:
            if 'B' in extracted_data[i][0] and curr_date == []:
                curr_date = [extracted_data[i][1]]
            elif 'B' in extracted_data[i][0] and curr_date != []:
                seqs_dict['date_seqs'].append(curr_date)
                curr_date = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_date == []:
                curr_date = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_date != []:
                curr_date.append(extracted_data[i][1])

        elif 'TOTAL' in extracted_data[i][0]:
            if 'B' in extracted_data[i][0] and curr_total == []:
                curr_total = [extracted_data[i][1]]
            elif 'B' in extracted_data[i][0] and curr_total != []:
                seqs_dict['total_seqs'].append(curr_total)
                curr_total = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_total == []:
                curr_total = [extracted_data[i][1]]
            elif 'I' in extracted_data[i][0] and curr_total != []:
                curr_total.append(extracted_data[i][1])

    # append last sequence
    seqs_dict['address_seqs'].append(curr_address)
    seqs_dict['company_seqs'].append(curr_company)
    seqs_dict['date_seqs'].append(curr_date)
    seqs_dict['total_seqs'].append(curr_total)

    # select longest sequence by characters for each entity
    entities_dict = {'address': max([' '.join(seqs) for seqs in seqs_dict['address_seqs']], key=len),
                     'company': max([' '.join(seqs) for seqs in seqs_dict['company_seqs']], key=len),
                     'date': max([' '.join(seqs) for seqs in seqs_dict['date_seqs']], key=len),
                     'total': max([' '.join(seqs) for seqs in seqs_dict['total_seqs']], key=len)}

    return entities_dict


def get_cell_coords_range_and_shape(sheet, cell):
    cell_range = f'{cell.coordinate}:{cell.coordinate}'
    width, height = 1, 1
    for rng in sheet.merged_cells.ranges:
        if cell.coordinate in rng:
            cell_range = rng.__str__()
            cell_root, cell_end = rng.__str__().split(':')
            width = column_index_from_string(cell_end[0]) - column_index_from_string(cell_root[0]) + 1
            height = int(''.join(cell_end[1:])) - int(''.join(cell_root[1:])) + 1

    return cell_range, width, height


def find_table_corners_by_two_values(corner_cell_values, sheet):
    # search for specific table values in the report
    # returns two (character, index) tuples of the cells with that values

    left_value, right_value = corner_cell_values[0], corner_cell_values[1]
    left_char, right_char = '', ''
    left_index, right_index = -1, -1

    # search left corner by scanning rows and columns
    # 'ABCDE' and 1-100 row range are magic numbers
    for c in 'ABCDE':
        for i in range(1, 100):
            cell_value = str(sheet[c + str(i)].value)
            if cell_value != str(None) and cell_value.strip().lower() == left_value:
                left_char = c
                left_index = i

    if left_index == -1:
        return left_char, left_index, right_char, right_index

    # search right corner by scanning columns from left-corner row
    for i in range(column_index_from_string(left_char) + 1, 100):  # 100 is a magic number
        cell_value = str(sheet[get_column_letter(i) + str(left_index)].value)
        if cell_value != str(None) and cell_value.strip().lower() == right_value:
            right_char = get_column_letter(i)
            right_index = left_index

    return left_char, left_index, right_char, right_index


def get_table_structure(sheet, start_row, end_row, start_col, end_col):
    subscription_structure = []
    seen_coords = set()
    # shape is row x columns, value is cell info tuple [(value, style, (range, width, height) )]
    for i in range(start_row, end_row+1):
        for j in range(start_col, end_col+1):
            cell_coord = get_column_letter(j) + str(i)
            if cell_coord not in seen_coords:
                cell = sheet[cell_coord]
                cell_range, width, height = get_cell_coords_range_and_shape(sheet, cell)
                cell_style = None
                if cell.has_style:
                    cell_style = {'font': copy(cell.font), 'border': copy(cell.border),
                                  'fill': copy(cell.fill), 'number_format': copy(cell.number_format),
                                  'protection': copy(cell.protection), 'alignment': copy(cell.alignment)}
                subscription_structure.append([cell.value, cell_style, (cell_range, width, height)])
                if width > 1 or height > 1:
                    coord_start, coord_end = cell_range.split(':')
                    start_char = coord_start[0]
                    start_idx = int(coord_start[1:])
                    for k in range(width):
                        next_col = column_index_from_string(start_char) + k
                        for l in range(height):
                            next_row = start_idx + l
                            seen_coords.add(f'{get_column_letter(next_col)}{next_row}')
                else:
                    seen_coords.add(cell_coord)
            else:
                continue

    return subscription_structure


def offset_table_structure(table_structure, row_offset, column_offset):
    offseted_structure = deepcopy(table_structure)
    for elem in offseted_structure:
        cell_range, width, height = elem[2]
        if width == 1 and height == 1:
            start_coord = cell_range.split(':')[0]
            start_char = start_coord[0]
            start_idx = int(start_coord[1:])
            new_char = get_column_letter(column_index_from_string(start_char) + column_offset)
            new_idx = start_idx + row_offset
            new_range = f'{new_char + str(new_idx)}:{new_char + str(new_idx)}'
            elem[2] = (new_range, width, height)
        else:
            start_coord = cell_range.split(':')[0]
            end_coord = cell_range.split(':')[1]
            start_char = start_coord[0]
            start_idx = int(start_coord[1:])
            end_char = end_coord[0]
            end_idx = int(end_coord[1:])
            new_start_char = get_column_letter(column_index_from_string(start_char) + column_offset)
            new_start_idx = start_idx + row_offset
            new_end_char = get_column_letter(column_index_from_string(end_char) + column_offset)
            new_end_idx = end_idx + row_offset
            new_range = f'{new_start_char + str(new_start_idx)}:{new_end_char + str(new_end_idx)}'
            elem[2] = (new_range, width, height)

    return offseted_structure


def range_contains(range_1, range_2):

    bound_1 = range_boundaries(range_1)
    bound_2 = range_boundaries(range_2)

    if bound_1[0] <= bound_2[0] and bound_1[1] <= bound_2[1] and bound_1[2] >= bound_2[2] and bound_1[3] >= bound_2[3]:
        return True
    else:
        return False


def clear_region_of_cells(sheet, cell_range):

    # Remove merge cells
    merge_cells_ranges = [x.coord for x in sheet.merged_cells.ranges]
    for merged_cells_range in merge_cells_ranges:
        if range_contains(cell_range, merged_cells_range):
            sheet.unmerge_cells(merged_cells_range)

    for column in cols_from_range(cell_range):
        for cell_coordinates in column:
            # Removing value
            sheet[cell_coordinates].value = None
            # Removing style (applying the style of cell A1)
            sheet[cell_coordinates]._style = sheet['B2']._style

    # Removing conditional formatting
    conditional_formattings = list(sheet.conditional_formatting._cf_rules.keys())
    for conditional_formatting in conditional_formattings:
        ranges_to_keep = [x for x in conditional_formatting.cells.ranges
                          if not range_contains(cell_range, x.coord)]

        if len(ranges_to_keep) != 0:
            conditional_formatting.cells.ranges = conditional_formatting.sqref.ranges = ranges_to_keep
        else:
            del sheet.conditional_formatting._cf_rules[conditional_formatting]

    # Removing data validation
    for validation in sheet.data_validations.dataValidation:
        for validation_range in validation.cells.ranges:
            if range_contains(cell_range, validation_range.coord):
                validation.cells.ranges.remove(validation_range)


def paste_table_structure_in_exel(sheet, table_structure):
    for elem in table_structure:
        cell_range, width, height = elem[2]

        cell_data_place = cell_range.split(':')[0]
        sheet[cell_data_place] = elem[0]
        cell_style = elem[1]
        if cell_style is not None:

            sheet[cell_data_place].font = cell_style['font']
            sheet[cell_data_place].border = cell_style['border']
            sheet[cell_data_place].fill = cell_style['fill']
            sheet[cell_data_place].number_format = cell_style['number_format']
            sheet[cell_data_place].protection = cell_style['protection']
            sheet[cell_data_place].alignment = cell_style['alignment']

        # if merged cell
        if width != 1 or height != 1:
            start_coord = cell_range.split(':')[0]
            end_coord = cell_range.split(':')[1]
            start_col = column_index_from_string(start_coord[0])
            start_idx = int(start_coord[1:])
            end_col = column_index_from_string(end_coord[0])
            end_idx = int(end_coord[1:])
            sheet.merge_cells(start_row=start_idx, start_column=start_col, end_row=end_idx, end_column=end_col)


def fill_exel_report(extracted_data, save_subs_structure, save_empty_table_row, report_filename, original_table_bottom_border=12):
    # extracted_data_list contains data for every receipt on the image
    book = openpyxl.open(report_filename, read_only=False)
    sheet = book.worksheets[0]

    save_empty_table_row = deepcopy(save_empty_table_row)
    save_subs_structure = deepcopy(save_subs_structure)

    # find indexes of the top corners of the table
    # values of the corner cells is hardcoded
    left_corner_value = 'Номер по порядку'.lower()
    right_corner_value = 'Расход (указать договор / ЦЗ)'.lower()

    left_char, left_index, right_char, right_index = find_table_corners_by_two_values((left_corner_value,
                                                                                      right_corner_value), sheet)

    table_numbers = []
    for i, row in enumerate(sheet):
        if number := sheet[left_char + str(left_index+4 + i)].value:  # 4 is magic number
            table_numbers.append(number)
        else:
            break

    table_number_top_border = min(table_numbers)
    table_number_bottom_border = max(table_numbers)

    # TODO: в будущем научиться детектить и заполнять номер чека, пока представим, что он всегда 59
    # copying contract cell to each row, because it's always the same for one report
    contract_cell_column_value = sheet[right_char + str(right_index+4)].value
    receipt_number_value = 59

    field_to_column_index_map = {'order_number': 0, 'date': 2, 'receipt_number': 4, 'company': 6,
                                 'total': 10, 'total_accepted': 13, 'debit': 16, 'contract': 19}

    entities_dict = parse_extracted_data(extracted_data)

    filled_row = False
    for i in range(table_number_bottom_border):

        date_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['date'])
        if sheet[date_column_char + str(left_index + i+4)].value == None:

            number_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['receipt_number'])
            company_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['company'])
            total_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['total'])
            contract_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['contract'])

            sheet[date_column_char + str(left_index + i+4)] = entities_dict['date']
            sheet[number_column_char + str(left_index + i+4)] = receipt_number_value
            sheet[company_column_char + str(left_index + i+4)] = entities_dict['company']
            sheet[total_column_char + str(left_index + i+4)] = entities_dict['total']
            sheet[contract_column_char + str(left_index + i+4)] = contract_cell_column_value

            filled_row = True
            break

    # add new row to the table and move subs structure
    if not filled_row:
        print('hello')
        # update new row value
        save_empty_table_row[1][0] = table_number_bottom_border + 1
        print(save_empty_table_row[1][0])

        # clear additional row under table and subs structure
        end_of_table = get_column_letter(column_index_from_string(right_char) + 4)

        range_to_clear = f'{left_char}{left_index+4 + table_number_bottom_border}:{end_of_table}{right_index+10 + table_number_bottom_border}'
        print(range_to_clear)
        clear_region_of_cells(sheet, range_to_clear)

        # offset and paste new row
        row_offset = table_number_bottom_border - original_table_bottom_border + 1
        print(row_offset)
        offset_row = offset_table_structure(save_empty_table_row, row_offset, 0)
        paste_table_structure_in_exel(sheet, offset_row)

        # changing row height for new row
        curr_last_row_index = left_index+4 + table_number_bottom_border - 1
        print(curr_last_row_index)
        sheet.row_dimensions[curr_last_row_index+1].height = sheet.row_dimensions[curr_last_row_index].height

        # offset and paste subscription structure
        save_subs_structure = offset_table_structure(save_subs_structure, row_offset, 0)
        paste_table_structure_in_exel(sheet, save_subs_structure)

        # past data
        number_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['receipt_number'])
        company_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['company'])
        total_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['total'])
        contract_column_char = get_column_letter(column_index_from_string(left_char) + field_to_column_index_map['contract'])

        sheet[date_column_char + str(left_index + table_number_bottom_border + 4)] = entities_dict['date']
        sheet[number_column_char + str(left_index + table_number_bottom_border + 4)] = receipt_number_value
        sheet[company_column_char + str(left_index + table_number_bottom_border + 4)] = entities_dict['company']
        sheet[total_column_char + str(left_index + table_number_bottom_border + 4)] = entities_dict['total']
        sheet[contract_column_char + str(left_index + table_number_bottom_border + 4)] = contract_cell_column_value

        # update border of table
        table_number_bottom_border += 1

    book.save(report_filename)
    book.close()


def fill_exel_report_good(extracted_data, clear_report_path, to_paste_report_path):
    book = openpyxl.open(clear_report_path, read_only=False)
    clear_sheet = book.worksheets[0]

    subscription_structure = get_table_structure(clear_sheet, 82, 86, 1, 25)
    empty_table_row = get_table_structure(clear_sheet, 79, 80, 1, 25)

    for data in extracted_data:
        fill_exel_report(data, subscription_structure, empty_table_row, original_table_bottom_border=12,
                         report_filename=to_paste_report_path)


if __name__ == '__main__':

    # just some testing
    nerv = [[('I-ADDRESS', '<s>'),
            ('B-ADDRESS', '115114'),
            ('I-ADDRESS', 'Mockba'),
            ('I-ADDRESS', 'Площадь'),
            ('I-ADDRESS', 'Павелечкого'),
            ('I-ADDRESS', 'вокзала'),
            ('I-ADDRESS', ''),
            ('I-ADDRESS', ','),
            ('I-ADDRESS', 'la'),
            ('B-ADDRESS', '115114,'),
            ('I-ADDRESS', 'r.Mockba,'),
            ('I-ADDRESS', 'nл.Павеле'),
            ('I-ADDRESS', 'ECC'),
            ('I-ADDRESS', ''),
            ('I-ADDRESS', '4A,'),
            ('B-DATE', ':31.10.2023'),
            ('I-ADDRESS', 'Aомодейово'),
            ('I-ADDRESS', '189'),
            ('I-ADDRESS', 'Павелeчkий'),
            ('B-DATE', '31.10.23'),
            ('B-TOTAL', ''),
            ('B-TOTAL', '=500.00'),
            ('B-TOTAL', '=3000.00'),
            ('I-COMPANY', 'OOO'),
            ('I-COMPANY', 'Слава'),
            ('I-COMPANY', 'Инфик'),
            ('B-COMPANY', 'Мышь')],

            [('B-ADDRESS', '115114,'),
             ('I-ADDRESS', 'Улица'),
             ('B-DATE', '15.02.1075'),
             ('I-ADDRESS', 'Пушкина'),
             ('I-ADDRESS', 'Дом'),
             ('I-ADDRESS', 'Колотушкина'),
             ('B-TOTAL', '=7500.00'),
             ('B-COMPANY', 'ООО'),
             ('I-COMPANY', 'Банк'),
             ('I-COMPANY', 'Приколов')],
            ]

    fill_exel_report_good(nerv, clear_report_path='exel_report_clear.xlsx', to_paste_report_path='exel_report_to_paste.xlsx')
